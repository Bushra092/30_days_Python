

from bs4 import BeautifulSoup
import openpyxl.workbook
import requests, openpyxl


excel = openpyxl.Workbook()
# print(excel.sheetnames)
sheet = excel.active
# print(sheet)
sheet.title = 'Top Rated Movies'
print(excel.sheetnames)

sheet.append(['Movie Rate','Movie name','Year', 'Rating'])


URL = 'https://www.imdb.com/chart/top/'
headers = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko)  Chrome/114.0.0.0 Safari/537.36"
}
try :
    source = requests.get(URL, headers=headers)
    source.raise_for_status()

    soup = BeautifulSoup(source.text, 'html.parser')
    movies = soup.find('ul', class_='ipc-metadata-list').find_all('li', class_="ipc-metadata-list-summary-item")
    

    for movie in movies:
        name = movie.find('h3',class_="ipc-title__text ipc-title__text--reduced" ).text  

        rank,name =  movie.find('h3',class_="ipc-title__text ipc-title__text--reduced").get_text(strip=True).split('.') 

        year = movie.find('span', class_="sc-dc48a950-8").text 
        rating= movie.find('span', class_="ipc-rating-star--rating").text
        print(rank,name,year,rating)
        sheet.append([rank,name,year,rating])

except Exception as e:
    print(e)

 

excel.save('Imdb Rate file')